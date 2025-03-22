import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def save_as_excel(results):
    print('엑셀 파일로 저장 중...')
    filename = "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "당근마켓 결과"
    ws.sheet_properties.tabColor = "1072BA"
    ws.append(["제목", "가격", "지역", "채팅", "조회", "등록일", "작성자", "설명"])

    default_font = Font(size=12)
    header_font = Font(size=14, bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    for idx, item in enumerate(results, start=2):
        title_cell = ws.cell(row=idx, column=1, value=item["title"])
        title_cell.hyperlink = item["link"]
        title_cell.style = "Hyperlink"
        
        ws.cell(row=idx, column=2, value=item["price"])
        ws.cell(row=idx, column=3, value=item["location"])
        ws.cell(row=idx, column=4, value=item["chatting"])
        ws.cell(row=idx, column=5, value=item["watching"])
        ws.cell(row=idx, column=6, value=item["time"])
        
        user_cell = ws.cell(row=idx, column=7, value=item["username"])
        user_cell.hyperlink = item["userlink"]
        user_cell.style = "Hyperlink"
        
        ws.cell(row=idx, column=8, value=item["description"])
        
        # Style and align each cell
        for col_idx in range(1, 9):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = default_font
            cell.alignment = center_align
            cell.border = thin_border

        # Adjust row height
        ws.row_dimensions[idx].height = 65

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max(max_length * 1.2, 12)  # 12 is the minimum width
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)