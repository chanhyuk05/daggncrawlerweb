import os
from io import BytesIO

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, Side


def save_as_excel(results):
    print('엑셀 파일로 저장 중...')
    filename = "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["제목", "가격", "상태", "링크"])

    default_font = Font(size=14)
    header_font = Font(size=16, bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.column_dimensions["A"].width = 40

    for idx, item in enumerate(results, start=2):
        # Insert values into columns B to E
        ws.cell(row=idx, column=1, value=item["title"])
        ws.cell(row=idx, column=2, value=item["price"])
        ws.cell(row=idx, column=3, value=item["status"])
        
        link_cell = ws.cell(row=idx, column=4, value="링크")
        link_cell.hyperlink = item["link"]
        link_cell.style = "Hyperlink"

        # Style and align each cell
        for col_idx in range(1, 5):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = default_font
            cell.alignment = center_align
            cell.border = thin_border
            if cell.value:
                current_width = ws.column_dimensions[cell.column_letter].width or 10
                new_width = max(current_width, len(str(cell.value)) * 1.2)
                ws.column_dimensions[cell.column_letter].width = new_width

        # Adjust row height
        ws.row_dimensions[idx].height = 65

    wb.save(filename)